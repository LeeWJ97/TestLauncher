import openpyxl
from openpyxl import load_workbook

class Excel():
    def __init__(self,filename,sheetname = 'Sheet1'):
        self.filename = filename
        self.sheetname = sheetname
        # 存储所有sheet
        self.wb = load_workbook(self.filename)
        # 存储sheet中的数据
        self.ws = self.wb[sheetname]

    def getallsheetname(self):
        return self.wb.get_sheet_names()


    def read(self,row,column):
        print(self.ws.cell(row,column).value)
        return self.ws.cell(row,column).value

    def write(self,row,column,value):
        # 存储sheet中的数据
        self.ws.cell(row = row, column = column, value = value)

    def save(self,filename=None):
        save_filename = self.filename if filename == None else filename
        self.wb.save(save_filename)


if __name__ == '__main__':
    r = Excel('data.xlsx','data')
    r.read(2,2)
    # r.write(11, 3,'测试')
    # r.write(11,4,'test2')
    r.save()
